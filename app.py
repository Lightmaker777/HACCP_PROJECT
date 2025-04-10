# app.py
from flask import Flask, render_template, request, redirect, url_for, session, Response, send_file, flash, jsonify
import bcrypt
import openpyxl 
from io import BytesIO
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from sqlalchemy import func
from datetime import datetime
from sqlalchemy.orm import validates
import os

# Initialize SQLAlchemy
db = SQLAlchemy()



# Define Models
class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False, index=True)
    password = db.Column(db.String(100), nullable=False)
    role = db.Column(db.String(20), default='mitarbeiter')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    products = db.relationship('Product', backref='added_by', lazy=True)
    confirmations = db.relationship('Confirmation', backref='user', lazy=True)
    
    @validates('username')
    def validate_username(self, key, username):
        if not username or len(username) < 3:
            raise ValueError("Username must be at least 3 characters long")
        return username

class ProductCategory(db.Model):
    __tablename__ = 'product_categories'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False, unique=True)
    min_temp = db.Column(db.Float)
    max_temp = db.Column(db.Float)
    risk_level = db.Column(db.String(20))
    
    # Relationship
    products = db.relationship('Product', backref='category', lazy=True)

class Product(db.Model):
    __tablename__ = 'produkte'
    id = db.Column(db.Integer, primary_key=True)
    produkt = db.Column(db.String(100), nullable=False)
    temperatur = db.Column(db.Float, nullable=False)
    lagerort = db.Column(db.String(100), nullable=False)
    status = db.Column(db.String(20), default='OK', index=True)
    risikostufe = db.Column(db.String(20), default='unbekannt', index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Foreign keys
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), index=True)
    category_id = db.Column(db.Integer, db.ForeignKey('product_categories.id'), index=True, nullable=True)
    
    @validates('temperatur')
    def validate_temperature(self, key, value):
        if not isinstance(value, (int, float)):
            raise ValueError("Temperature must be a number")
        return value
    
    @validates('produkt')
    def validate_product(self, key, value):
        if not value or len(value) < 2:
            raise ValueError("Product name must be at least 2 characters")
        return value

class SecurityCheck(db.Model):
    __tablename__ = 'sicherheit'
    id = db.Column(db.Integer, primary_key=True)
    faktor = db.Column(db.String(100), nullable=False)
    überprüfung = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Foreign key
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), index=True)

class Confirmation(db.Model):
    __tablename__ = 'confirmations'
    id = db.Column(db.Integer, primary_key=True)
    employee_name = db.Column(db.String(100), nullable=False)
    confirmation_date = db.Column(db.String(10), nullable=False, index=True)
    employee_number = db.Column(db.String(50), nullable=False, index=True)
    signature = db.Column(db.Text, nullable=False)
    instructor_name = db.Column(db.String(100))
    instruction_version = db.Column(db.String(50))
    employee_role = db.Column(db.String(50))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Foreign key
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), index=True)

# Initialize the Flask application
app = Flask(__name__)
# Datenbank-Verbindung
app.config.update(
    SECRET_KEY='your_secret_key',  # Den Schlüssel solltest du sicher verwahren
    SQLALCHEMY_DATABASE_URI=os.environ.get('DATABASE_URL', 'sqlite:///haccp.db'),  # Hier greifen wir auf die Umgebungsvariable zu
    SQLALCHEMY_TRACK_MODIFICATIONS=False,
    SQLALCHEMY_POOL_SIZE=10,
    SQLALCHEMY_POOL_TIMEOUT=30,
    SQLALCHEMY_POOL_RECYCLE=1800,
    SQLALCHEMY_MAX_OVERFLOW=20
)

# Initialize SQLAlchemy with the app
db.init_app(app)
migrate = Migrate(app, db)

# Password handling functions
def hash_password(password):
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())

def check_password(stored_password, password):
    return bcrypt.checkpw(password.encode('utf-8'), stored_password)

# Database initialization
def initialize():
    with app.app_context():
        db.create_all()  # Creates tables if they don't exist
        create_default_admin()
        create_product_categories()

def create_default_admin():
    admin = User.query.filter_by(username='admin').first()
    if not admin:
        hashed_password = hash_password('admin123')
        admin = User(username='admin', password=hashed_password, role='admin')
        db.session.add(admin)
        db.session.commit()
        
def create_product_categories():
    categories = [
        {"name": "Fleisch", "min_temp": 2, "max_temp": 7, "risk_level": "hoch"},
        {"name": "Milch", "min_temp": 0, "max_temp": 4, "risk_level": "hoch"},
        {"name": "Gemüse", "min_temp": 4, "max_temp": 12, "risk_level": "mittel"},
        {"name": "Honig", "min_temp": -99, "max_temp": 99, "risk_level": "niedrig"}
    ]
    
    for cat_data in categories:
        if not ProductCategory.query.filter_by(name=cat_data["name"]).first():
            category = ProductCategory(
                name=cat_data["name"],
                min_temp=cat_data["min_temp"],
                max_temp=cat_data["max_temp"],
                risk_level=cat_data["risk_level"]
            )
            db.session.add(category)
    
    db.session.commit()

# Authentication handling
@app.route('/register', methods=['GET', 'POST'])
def register():
    if 'user' in session:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        # Check if username exists
        existing_user = User.query.filter_by(username=username).first()
        if existing_user:
            flash("Benutzername ist bereits vergeben.", "danger")
            return render_template('register.html')

        try:
            # Create new user
            hashed_password = hash_password(password)
            new_user = User(username=username, password=hashed_password)
            db.session.add(new_user)
            db.session.commit()
            
            flash("Registrierung erfolgreich! Sie können sich jetzt anmelden.", "success")
            return redirect(url_for('login'))
        except Exception as e:
            db.session.rollback()
            flash(f"Fehler bei der Registrierung: {str(e)}", "danger")
    
    return render_template('register.html')

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        user = User.query.filter_by(username=username).first()
        
        if user and check_password(user.password, password):
            session['user'] = user.username
            session['user_id'] = user.id
            session['role'] = user.role
            return redirect(url_for('dashboard'))
        else:
            flash("Ungültige Anmeldedaten!", "danger")
    
    return render_template('login.html')

@app.route("/dashboard")
def dashboard():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    # Get some stats for the dashboard
    product_count = Product.query.count()
    warning_count = Product.query.filter_by(status="WARNUNG").count()
    confirmation_count = Confirmation.query.count()
    
    return render_template('dashboard.html', 
                          product_count=product_count, 
                          warning_count=warning_count,
                          confirmation_count=confirmation_count)

@app.route("/logout")
def logout():
    session.pop('user', None)
    session.pop('user_id', None)
    session.pop('role', None)
    return redirect(url_for('login'))

# Route protection middleware
@app.before_request
def require_login_and_admin():
    public_routes = ['login', 'register', 'static']
    
    if request.endpoint in public_routes:
        return None

    if 'user' not in session:
        return redirect(url_for('login'))
    else:
        admin_routes = ['admin_dashboard', 'admin_settings']
        
        if request.endpoint in admin_routes and session.get('role') != 'admin':
            flash("Sie haben keine Berechtigung für diese Seite.", "danger")
            return redirect(url_for('dashboard'))

# Main product routes
@app.route("/", methods=["GET", "POST"])
def index():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    if request.method == "POST":
        produkt = request.form["produkt"]
        temperatur = float(request.form["temperatur"])
        lagerort = request.form["lagerort"]
        
        # Get the product category
        category = ProductCategory.query.filter_by(name=produkt).first()
        
        status = "OK"
        risikostufe = "unbekannt"
        
        if category:
            risikostufe = category.risk_level
            if temperatur < category.min_temp or temperatur > category.max_temp:
                status = "WARNUNG"
        
        try:
            # Create new product
            new_product = Product(
                produkt=produkt,
                temperatur=temperatur,
                lagerort=lagerort,
                status=status,
                risikostufe=risikostufe,
                user_id=session.get('user_id'),
                category_id=category.id if category else None
            )
            db.session.add(new_product)
            db.session.commit()
            
            # Redirect to confirmation page
            return redirect(url_for('confirmation', produkt=produkt, temperatur=temperatur, status=status, risikostufe=risikostufe))
        except Exception as e:
            db.session.rollback()
            flash(f"Fehler beim Speichern: {str(e)}", "danger")
    
    # Get product categories for dropdown
    categories = ProductCategory.query.all()
    return render_template("index.html", categories=categories)

@app.route("/confirmation", methods=["GET", "POST"])
def confirmation():
    produkt = request.args.get('produkt')
    temperatur = request.args.get('temperatur')
    status = request.args.get('status')
    risikostufe = request.args.get('risikostufe')

    if request.method == "POST":
        if request.is_json:
            data = request.get_json()
            
            try:
                # Create new confirmation
                new_confirmation = Confirmation(
                    employee_name=data['employee_name'],
                    confirmation_date=data['confirmation_date'],
                    signature=data['signature'],
                    employee_number=data['employee_number'],
                    instructor_name=data['instructor_name'],
                    instruction_version=data['instruction_version'],
                    employee_role=data['employee_role'],
                    user_id=session.get('user_id')
                )
                db.session.add(new_confirmation)
                db.session.commit()
                
                return jsonify({"message": "Bestätigung gespeichert"}), 200
            except Exception as e:
                db.session.rollback()
                return jsonify({"error": str(e)}), 400
        else:
            return jsonify({"error": "Erwarte JSON-Daten"}), 400

    return render_template("confirmation.html", 
                          produkt=produkt, 
                          temperatur=temperatur, 
                          status=status, 
                          risikostufe=risikostufe)

@app.route("/confirmations")
def confirmations():
    if 'user' not in session:
        return redirect(url_for('login'))

    confirmations = Confirmation.query.order_by(Confirmation.created_at.desc()).all()
    return render_template("confirmations.html", rows=confirmations)

@app.route("/produkte", methods=["GET", "POST"])
def produkte():
    if 'user' not in session:
        return redirect(url_for('login'))

    products = Product.query.order_by(Product.created_at.desc()).all()
    
    if request.method == "POST":
        produkt = request.form["produkt"]
        temperatur = float(request.form["temperatur"])
        lagerort = request.form["lagerort"]
        
        # Get the product category
        category = ProductCategory.query.filter_by(name=produkt).first()
        
        status = "OK"
        risikostufe = "unbekannt"
        
        if category:
            risikostufe = category.risk_level
            if temperatur < category.min_temp or temperatur > category.max_temp:
                status = "WARNUNG"
        
        try:
            # Create new product
            new_product = Product(
                produkt=produkt,
                temperatur=temperatur,
                lagerort=lagerort,
                status=status,
                risikostufe=risikostufe,
                user_id=session.get('user_id'),
                category_id=category.id if category else None
            )
            db.session.add(new_product)
            db.session.commit()
            
            flash(f"{produkt} erfolgreich validiert!", "success")
            return redirect(url_for('produkte'))
        except Exception as e:
            db.session.rollback()
            flash(f"Fehler beim Speichern: {str(e)}", "danger")

    # Get product categories for dropdown
    categories = ProductCategory.query.all()
    return render_template("produkte.html", rows=products, categories=categories)

@app.route("/produkte_validierung", methods=["GET", "POST"])
def produkte_validierung():
    if 'user' not in session:
        return redirect(url_for('login'))

    if request.method == "POST":
        produkt = request.form["produkt"]
        temperatur = float(request.form["temperatur"])
        lagerort = request.form["lagerort"]
        
        # Get the product category
        category = ProductCategory.query.filter_by(name=produkt).first()
        
        status = "OK"
        risikostufe = "unbekannt"
        
        if category:
            risikostufe = category.risk_level
            if temperatur < category.min_temp or temperatur > category.max_temp:
                status = "WARNUNG"
        
        try:
            # Create new product
            new_product = Product(
                produkt=produkt,
                temperatur=temperatur,
                lagerort=lagerort,
                status=status,
                risikostufe=risikostufe,
                user_id=session.get('user_id'),
                category_id=category.id if category else None
            )
            db.session.add(new_product)
            db.session.commit()
            
            flash(f"Produkt {produkt} erfolgreich validiert!", "success")
            return redirect(url_for('produkte_validierung'))
        except Exception as e:
            db.session.rollback()
            flash(f"Fehler beim Speichern: {str(e)}", "danger")

    # Get product categories for dropdown
    categories = ProductCategory.query.all()
    return render_template("produkte_validierung.html", categories=categories)

@app.route("/sicherheit", methods=["GET", "POST"])
def sicherheit():
    if 'user' not in session:
        return redirect(url_for('login'))

    if request.method == "POST":
        sicherheitsfaktor = request.form["sicherheitsfaktor"]
        überprüfung = request.form["überprüfung"]

        try:
            # Create new security check
            new_check = SecurityCheck(
                faktor=sicherheitsfaktor,
                überprüfung=überprüfung,
                user_id=session.get('user_id')
            )
            db.session.add(new_check)
            db.session.commit()
            
            flash(f"Sicherheitsüberprüfung für {sicherheitsfaktor} abgeschlossen!", "success")
            return redirect(url_for('sicherheit'))
        except Exception as e:
            db.session.rollback()
            flash(f"Fehler beim Speichern: {str(e)}", "danger")

    return render_template("sicherheit.html")

# Export functionalities
@app.route("/export")
def export_csv():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    products = Product.query.all()
    
    # Create CSV content
    output = "ID,Produkt,Temperatur,Lagerort,Status,Risikostufe\n"
    for product in products:
        output += f"{product.id},{product.produkt},{product.temperatur},{product.lagerort},{product.status},{product.risikostufe}\n"

    return Response(
        output,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment;filename=produkte.csv"}
    )

@app.route("/export_excel")
def export_excel():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    products = Product.query.all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ID", "Produkt", "Temperatur", "Lagerort", "Status", "Risikostufe"])

    for product in products:
        ws.append([
            product.id, 
            product.produkt, 
            product.temperatur, 
            product.lagerort, 
            product.status, 
            product.risikostufe
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output, 
        as_attachment=True, 
        download_name="produkte.xlsx", 
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Statistics and reporting
@app.route("/statistiken")
def statistiken():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    # Efficient aggregation using SQLAlchemy
    avg_temp = db.session.query(func.avg(Product.temperatur)).scalar() or 0
    
    warning_count = Product.query.filter_by(status="WARNUNG").count()
    total_count = Product.query.count()
    
    # Products by risk level
    risk_stats = db.session.query(
        Product.risikostufe,
        func.count(Product.id).label('count')
    ).group_by(Product.risikostufe).all()
    
    # Products by category
    category_stats = db.session.query(
        ProductCategory.name,
        func.count(Product.id).label('count')
    ).join(Product, Product.category_id == ProductCategory.id, isouter=True)\
     .group_by(ProductCategory.name).all()
    
    return render_template("statistiken.html", 
                          avg_temp=avg_temp, 
                          warning_count=warning_count, 
                          total_count=total_count,
                          risk_stats=risk_stats,
                          category_stats=category_stats)

# Batch operations example
@app.route("/admin/bulk_import", methods=["GET", "POST"])
def bulk_import():
    if 'user' not in session or session.get('role') != 'admin':
        flash("Sie haben keine Berechtigung für diese Seite.", "danger")
        return redirect(url_for('dashboard'))
    
    if request.method == "POST":
        # Handle file upload and processing here
        if 'file' not in request.files:
            flash('Keine Datei ausgewählt', 'danger')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('Keine Datei ausgewählt', 'danger')
            return redirect(request.url)
        
        try:
            # Read the Excel file
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            products = []
            # Skip header row
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) >= 4:  # Ensure row has enough data
                    product_name = row[0]
                    temperature = float(row[1])
                    location = row[2]
                    
                    # Get product category
                    category = ProductCategory.query.filter_by(name=product_name).first()
                    
                    status = "OK"
                    risk_level = "unbekannt"
                    
                    if category:
                        risk_level = category.risk_level
                        if temperature < category.min_temp or temperature > category.max_temp:
                            status = "WARNUNG"
                    
                    # Create product object
                    new_product = Product(
                        produkt=product_name,
                        temperatur=temperature,
                        lagerort=location,
                        status=status,
                        risikostufe=risk_level,
                        user_id=session.get('user_id'),
                        category_id=category.id if category else None
                    )
                    products.append(new_product)
            
            # Bulk insert
            db.session.bulk_save_objects(products)
            db.session.commit()
            
            flash(f"{len(products)} Produkte erfolgreich importiert!", "success")
            return redirect(url_for('produkte'))
        except Exception as e:
            db.session.rollback()
            flash(f"Fehler beim Import: {str(e)}", "danger")
    
    return render_template("bulk_import.html")

# Admin functionality
@app.route("/admin/categories", methods=["GET", "POST"])
def manage_categories():
    if 'user' not in session or session.get('role') != 'admin':
        flash("Sie haben keine Berechtigung für diese Seite.", "danger")
        return redirect(url_for('dashboard'))
    
    if request.method == "POST":
        name = request.form["name"]
        min_temp = float(request.form["min_temp"])
        max_temp = float(request.form["max_temp"])
        risk_level = request.form["risk_level"]
        
        try:
            # Create new category
            new_category = ProductCategory(
                name=name,
                min_temp=min_temp,
                max_temp=max_temp,
                risk_level=risk_level
            )
            db.session.add(new_category)
            db.session.commit()
            
            flash(f"Kategorie {name} erfolgreich erstellt!", "success")
            return redirect(url_for('manage_categories'))
        except Exception as e:
            db.session.rollback()
            flash(f"Fehler beim Speichern: {str(e)}", "danger")
    
    categories = ProductCategory.query.all()
    return render_template("manage_categories.html", categories=categories)

if __name__ == "__main__":
    # Initialize database before running the app
    with app.app_context():
        initialize()
    app.run(debug=True)